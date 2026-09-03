# -*- coding: utf-8 -*-
"""Khởi tạo MỘT chiến dịch mới từ template chuẩn — khâu ① của quy trình 7 khâu.

    new_campaign.py --code 02_Ten_Chien_Dich [--meta meta.json] [--instance <ten>]
                    [--station <path>] [--dry-run]

NGUYÊN TẮC CỐT LÕI — COPY, KHÔNG DỰNG LẠI:
    Workbook được tạo bằng `shutil.copy2` từ `templates/CAMPAIGN_TEMPLATE.xlsx`, sau đó
    mới sửa GIÁ TRỊ ô. KHÔNG bao giờ dựng workbook mới bằng openpyxl.Workbook().

    Vì sao: dựng lại thì mất sạch màu, độ rộng cột, freeze pane, định dạng số — tức mất
    đúng phần làm file này đọc được bằng mắt người. Template là hợp đồng hình thức, không
    chỉ là danh sách cột. `build_workbook.py` (5 sheet, dựng từ spec) làm ngược lại và vì
    thế đã bị thay bằng script này.

DỮ LIỆU SỐNG Ở STATION, KHÔNG Ở REPO:
    Repo là XƯỞNG — chỉ chứa quy trình, skill, template, dữ liệu mẫu. Chiến dịch thật
    nằm ở STATION của người dùng. Thứ tự phân giải (dừng ở cái đầu tiên tìm thấy):
      1. --station tường minh
      2. thư mục làm việc hiện tại nếu có dấu hiệu dự án (.marketing-studio/ hoặc instance.yml)
         -> làm việc trong folder riêng thì asset sinh THẲNG vào đó
      3. biến môi trường MARKETING_STUDIO_DATA
      4. ~/.marketing
"""
from __future__ import annotations

import argparse
import json
import os
import shutil
import sys
from datetime import date
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Thiếu openpyxl. Cần: pip install openpyxl", file=sys.stderr)
    sys.exit(2)

REPO = Path(__file__).resolve().parents[2]
TPL_XLSX = REPO / "templates" / "CAMPAIGN_TEMPLATE.xlsx"
TPL_MD = REPO / "templates" / "CAMPAIGN_TEMPLATE.md"
SHEETS = ("Campaign", "Content", "Post", "_Legend")
MARKERS = (".marketing-studio", "instance.yml")

# Ánh xạ khoá trong campaign_meta.json -> tên trường ở sheet Campaign.
# Chỉ ánh xạ khi TÊN KHÁC NHAU; khoá trùng tên trường thì tự khớp.
META_ALIAS = {
    "objective": "campaign_goal",
    "pillar": "content_pillar",
}


def resolve_station(explicit: str | None) -> tuple[Path, str]:
    """(đường dẫn STATION, lý do chọn) — xem docstring đầu file."""
    if explicit:
        return Path(explicit).expanduser(), "--station tường minh"
    cwd = Path.cwd()
    for d in (cwd, *cwd.parents):
        if any((d / m).exists() for m in MARKERS):
            return d, f"thư mục dự án (thấy {MARKERS[0]}/{MARKERS[1]} tại {d})"
        if d == d.parent:
            break
    env = os.environ.get("MARKETING_STUDIO_DATA")
    if env:
        return Path(env).expanduser(), "biến MARKETING_STUDIO_DATA"
    return Path.home() / ".marketing", "mặc định ~/.marketing"


def load_meta(p: str | None) -> dict:
    if not p:
        return {}
    d = json.loads(Path(p).read_text(encoding="utf-8"))
    return {META_ALIAS.get(k, k): v for k, v in d.items()}


def fill_workbook(xlsx: Path, code: str, meta: dict) -> list[str]:
    """Dọn dữ liệu mẫu + đổ sheet Campaign. Trả về cảnh báo (nếu có)."""
    wb = load_workbook(xlsx)
    warn: list[str] = []

    missing = [s for s in SHEETS if s not in wb.sheetnames]
    if missing:
        raise SystemExit(f"Template hỏng: thiếu sheet {missing}. Dừng, không tạo file lỗi.")

    ws = wb["Campaign"]
    # Bảng tên-trường -> dòng. Bỏ qua dòng tiêu đề nhóm (cột B trống trong template gốc
    # VÀ tên viết HOA) — chúng là nhãn nhóm, không phải trường.
    rows: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = ws.cell(r, 1).value
        if isinstance(k, str) and k.strip() and not (k.isupper() and " " in k):
            rows[k.strip()] = r

    for f, r in rows.items():
        # Dọn TRƯỚC: giá trị mẫu của template không được lẫn sang chiến dịch mới.
        ws.cell(r, 2).value = None

    # meta được phép ghi đè phần lớn trường, NHƯNG KHÔNG được đè campaign_code:
    # mã này quyết định TÊN THƯ MỤC và TÊN FILE, nên ô trong sheet phải khớp --code.
    # Để meta đè thì workbook khai một mã còn thư mục mang mã khác — lệch âm thầm,
    # và mọi script tra theo mã sẽ không tìm thấy. (bug bắt được lúc tự test 03/09)
    defaults = {"status": "active", "created": date.today().isoformat()}
    values = {**defaults, **{k: v for k, v in meta.items() if k in rows}}
    values["campaign_code"] = code           # LUÔN thắng, không thương lượng
    for f, v in values.items():
        ws.cell(rows[f], 2).value = v

    unknown = [k for k in meta if k not in rows]
    if unknown:
        warn.append(f"meta có {len(unknown)} khoá không phải trường Campaign, đã BỎ QUA: {unknown}")
    empty = [f for f in rows if ws.cell(rows[f], 2).value in (None, "")]
    if empty:
        warn.append(f"{len(empty)}/{len(rows)} trường Campaign còn TRỐNG — điền nốt trước khi sang khâu ②: {empty}")

    # Content/Post: xoá GIÁ TRỊ, không xoá dòng — xoá dòng làm hỏng định dạng còn lại.
    for s in ("Content", "Post"):
        t = wb[s]
        for r in range(2, t.max_row + 1):
            for c in range(1, t.max_column + 1):
                t.cell(r, c).value = None

    wb.save(xlsx)
    return warn


def main() -> int:
    ap = argparse.ArgumentParser(description="Khởi tạo chiến dịch mới từ template chuẩn.")
    ap.add_argument("--code", required=True, help="mã chiến dịch, vd 02_Ten_Chien_Dich")
    ap.add_argument("--meta", help="campaign_meta.json để đổ sẵn sheet Campaign")
    ap.add_argument("--instance", default="default", help="tên instance trong STATION")
    ap.add_argument("--station", help="ghi đè đường dẫn STATION")
    ap.add_argument("--dry-run", action="store_true", help="chỉ in ra, không ghi gì")
    a = ap.parse_args()

    for t in (TPL_XLSX, TPL_MD):
        if not t.is_file():
            print(f"LỖI: thiếu template {t}", file=sys.stderr)
            return 2

    station, why = resolve_station(a.station)
    dest = station / "instances" / a.instance / "02_campaigns" / a.code
    print(f"STATION : {station}   ({why})")
    print(f"Đích    : {dest}")

    if dest.exists() and any(dest.iterdir()):
        print(f"LỖI: {dest} đã tồn tại và KHÔNG rỗng. Dừng — không ghi đè chiến dịch có sẵn.",
              file=sys.stderr)
        return 1
    if a.dry_run:
        print("(dry-run — không ghi gì)")
        return 0

    (dest / "assets").mkdir(parents=True, exist_ok=True)
    xlsx, md = dest / f"{a.code}.xlsx", dest / f"{a.code}.md"
    shutil.copy2(TPL_XLSX, xlsx)   # copy2 giữ nguyên nội dung + mtime; định dạng nguyên vẹn
    shutil.copy2(TPL_MD, md)

    warn = fill_workbook(xlsx, a.code, load_meta(a.meta))

    wb = load_workbook(xlsx)
    print(f"OK      : {xlsx}")
    print(f"OK      : {md}")
    print(f"Sheet   : {wb.sheetnames}")
    for w in warn:
        print(f"CẢNH BÁO: {w}")
    print("Tiếp    : điền nốt sheet Campaign rồi sang khâu ② (workflows/02_plan_content.md)")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.exit(main())

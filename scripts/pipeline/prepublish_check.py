# -*- coding: utf-8 -*-
r"""prepublish_check.py — GUARDRAIL trước khi publish: liệt kê & kiểm đủ ASSET + LINK.

Trả exit 0 nếu đủ mục bắt buộc của --mode; exit 2 + liệt kê thiếu nếu chưa đủ.
In bảng checklist rõ ràng để người/agent thấy bài đã sẵn sàng chưa.

Mode:
  assets : kiểm file trong folder bài (blog/fb/yt-desc/thumbnail/video/atlas.html/meta...).
  links  : kiểm Sheet Result của bài đã có blog_url + youtube_url (cổng chặn TRƯỚC khi đăng FB).
  all    : cả hai.

CLI:
  python prepublish_check.py --folder <FOLDER> --mode assets
  python prepublish_check.py --xlsx <campaign.xlsx> --post-id ID --mode links
  python prepublish_check.py --folder <FOLDER> --xlsx <campaign.xlsx> --post-id ID --mode all
"""
from __future__ import annotations
import argparse, json, os, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
import post_paths as PP  # noqa: E402

# (khoá trong post_paths.LAYOUT, bắt buộc?)
# comment.txt và infographic.* nay là BẮT BUỘC: luật 04/09 đòi thân bài Facebook 0 URL
# (nên link phải nằm ở comment) và đòi ảnh tóm tắt kèm sidecar prompt.
ASSETS = [(PP.LAYOUT[k], b) for k, b in [
    ("meta", True), ("research", True), ("content", False),
    ("blog", True), ("atlas_html", True), ("audio", False),
    ("yt_desc", True), ("yt_thumb", True), ("yt_video", True),
    ("fb_post", True), ("fb_comment", True),
    ("fb_image", True), ("fb_prompt", True),
]]
LINKS_REQUIRED = ["blog_url", "youtube_url"]  # phải đủ trước khi đăng FB


def _result_row(xlsx, post_id):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True)
    if "Result" not in wb.sheetnames:
        return {}
    ws = wb["Result"]
    hdr = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {hdr[i]: r[i] for i in range(len(hdr)) if i < len(r)}
        if str(d.get("post_id") or "") == post_id:
            return d
    return {}


def check_assets(folder):
    rows, missing = [], []
    for name, req in ASSETS:
        p = os.path.join(folder, name)
        ok = os.path.isfile(p) and os.path.getsize(p) > 0
        rows.append((name, req, ok))
        if req and not ok:
            missing.append(name)
    return rows, missing


def check_links(xlsx, post_id):
    d = _result_row(xlsx, post_id)
    rows, missing = [], []
    for k in LINKS_REQUIRED + ["fb_post_id", "fb_permalink"]:
        v = str(d.get(k) or "").strip()
        req = k in LINKS_REQUIRED
        ok = bool(v)
        rows.append((k, req, ok, v))
        if req and not ok:
            missing.append(k)
    return rows, missing


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--folder")
    ap.add_argument("--xlsx")
    ap.add_argument("--post-id")
    ap.add_argument("--mode", choices=["assets", "links", "all"], default="all")
    a = ap.parse_args(argv)

    missing_total = []
    print("=== PRE-PUBLISH CHECKLIST ===")
    if a.mode in ("assets", "all"):
        if not a.folder:
            print("  (assets) THIẾU --folder"); missing_total.append("--folder")
        else:
            rows, miss = check_assets(a.folder)
            print("ASSET (file):")
            for name, req, ok in rows:
                print(f"  [{'x' if ok else ' '}] {name:18} {'(bắt buộc)' if req else '(tùy chọn)'}")
            missing_total += miss
    if a.mode in ("links", "all"):
        if not (a.xlsx and a.post_id):
            print("  (links) THIẾU --xlsx/--post-id"); missing_total.append("--xlsx/--post-id")
        else:
            rows, miss = check_links(a.xlsx, a.post_id)
            print("LINK (Sheet Result) — phải đủ TRƯỚC khi đăng FB:")
            for k, req, ok, v in rows:
                show = (v[:48] + "…") if len(v) > 49 else v
                print(f"  [{'x' if ok else ' '}] {k:14} {'(bắt buộc)' if req else '(sau FB)'}  {show}")
            missing_total += miss

    if missing_total:
        print("RESULT: THIẾU -> " + ", ".join(missing_total))
        print("NOT_READY")
        return 2
    print("RESULT: đủ mục bắt buộc.")
    print("READY")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    raise SystemExit(main())

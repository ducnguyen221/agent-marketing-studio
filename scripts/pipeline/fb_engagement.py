# -*- coding: utf-8 -*-
"""fb_engagement.py — quét chỉ số tương tác Facebook cho 1 campaign Excel.

Đọc sheet **Result** lấy mọi (post_id, fb_post_id). Với mỗi fb_post_id còn sống:
  GET {GRAPH}/{fb_post_id}?fields=permalink_url,shares,
        likes.summary(true),comments.summary(true),reactions.summary(true)
  -> likes / comments / reactions / shares + permalink (ĐÃ verify lấy được).
  reach/impressions: thử {GRAPH}/{fb_post_id}/insights?metric=... — BEST-EFFORT,
  lỗi (400/quyền) thì để trống, KHÔNG làm hỏng cả script.

Upsert vào sheet **Engagement** qua CLI `tobi_excel.py upsert-engagement --json`
(theo PIPELINE_CONTRACT — KHÔNG phụ thuộc nội bộ tobi_excel.py vì tầng Excel
đang được agent khác viết lại).

Chạy bằng PYTHON HỆ THỐNG (có `requests`). KHÔNG in token ra log/stdout.

  PY fb_engagement.py --campaign-xlsx <campaign.xlsx> [--fb-config <json>]

In tóm tắt số bài cập nhật; dòng cuối: `OK <campaign-xlsx>`.
"""
from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
import tempfile
import time

import requests

GRAPH = "https://graph.facebook.com/v21.0"  # version giống post_facebook.py
DEFAULT_FB_CONFIG = os.environ.get("FB_CONFIG") or os.path.join(
    os.path.expanduser("~"), ".video", "tobi", "facebook_config.json")
_HERE = os.path.dirname(os.path.abspath(__file__))
TOBI_EXCEL = os.path.join(_HERE, "tobi_excel.py")

# fields ĐÃ verify lấy được counts qua Graph (object đơn lẻ).
_FIELDS = ("permalink_url,shares,"
           "likes.summary(true),comments.summary(true),reactions.summary(true)")
# metric thử cho insights (best-effort; nhiều khả năng 400 nếu thiếu read_insights).
_INSIGHT_METRICS = "post_impressions_unique,post_impressions"


# ---------------------------------------------------------------- config FB

def _load_fb_config(path):
    if not os.path.isfile(path):
        sys.exit(f"Thiếu facebook_config.json: {path}")
    with open(path, encoding="utf-8-sig") as f:
        c = json.load(f)
    if not c.get("page_token"):
        sys.exit(f"Config thiếu page_token: {path}")
    return c


# ---------------------------------------------------------------- đọc Excel (read-only, schema-tolerant)

def _read_result_rows(xlsx_path):
    """Đọc sheet Result -> list dict theo header thực tế của file.

    Đọc header động (không hard-code thứ tự cột) để bền với việc tầng Excel
    đang được viết lại. Trả [] nếu chưa có file hoặc chưa có sheet Result.
    """
    if not os.path.isfile(xlsx_path):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if "Result" not in wb.sheetnames:
            return []
        ws = wb["Result"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {}
        empty = True
        for i, col in enumerate(header):
            if not col:
                continue
            v = r[i] if i < len(r) else None
            if v not in (None, ""):
                empty = False
            rec[col] = v
        if not empty:
            out.append(rec)
    return out


# ---------------------------------------------------------------- gọi Graph

def _summary_count(node, key):
    """Lấy summary.total_count cho likes/comments/reactions; None nếu thiếu."""
    sub = node.get(key)
    if isinstance(sub, dict):
        summ = sub.get("summary")
        if isinstance(summ, dict) and summ.get("total_count") is not None:
            return summ["total_count"]
    return None


def _shares_count(node):
    sh = node.get("shares")
    if isinstance(sh, dict):
        return sh.get("count")
    return None


def fetch_engagement(fb_post_id, token, session):
    """GET object đơn lẻ -> dict các chỉ số cơ bản (đã verify) + permalink."""
    r = session.get(f"{GRAPH}/{fb_post_id}",
                    params={"fields": _FIELDS, "access_token": token}, timeout=30)
    r.raise_for_status()
    d = r.json()
    return {
        "fb_permalink": d.get("permalink_url") or "",
        "likes": _summary_count(d, "likes"),
        "comments": _summary_count(d, "comments"),
        "reactions": _summary_count(d, "reactions"),
        "shares": _shares_count(d),
    }


def fetch_insights_best_effort(fb_post_id, token, session):
    """Thử insights reach/impressions. BEST-EFFORT: trả {} nếu lỗi (400/quyền).

    KHÔNG raise — reach/impressions để trống nếu không lấy được.
    """
    try:
        r = session.get(f"{GRAPH}/{fb_post_id}/insights",
                        params={"metric": _INSIGHT_METRICS, "access_token": token},
                        timeout=30)
        if r.status_code != 200:
            return {}
        data = r.json().get("data", [])
        out = {}
        for m in data:
            name = m.get("name")
            vals = m.get("values") or []
            val = vals[0].get("value") if vals and isinstance(vals[0], dict) else None
            if name in ("post_impressions_unique",) and val is not None:
                out["reach"] = val
            elif name in ("post_impressions",) and val is not None:
                out["impressions"] = val
        return out
    except Exception:
        return {}


# ---------------------------------------------------------------- ghi Engagement qua CLI

def _upsert_engagement(xlsx_path, row):
    """Ghi 1 hàng sheet Engagement qua CLI `tobi_excel.py upsert-engagement --json`.

    Dùng CLI theo CONTRACT (không import nội bộ). Tạo file JSON tạm.
    """
    fd, tmp = tempfile.mkstemp(suffix=".json", prefix="eng_")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(row, f, ensure_ascii=False)
        proc = subprocess.run(
            [sys.executable, TOBI_EXCEL, "upsert-engagement",
             "--path", xlsx_path, "--json", tmp],
            capture_output=True, text=True, encoding="utf-8")
        if proc.returncode != 0:
            err = (proc.stderr or proc.stdout or "").strip().splitlines()
            tail = err[-1] if err else f"exit {proc.returncode}"
            raise RuntimeError(f"upsert-engagement lỗi: {tail}")
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def run(campaign_xlsx, fb_config):
    c = _load_fb_config(fb_config)
    token = c["page_token"]
    rows = _read_result_rows(campaign_xlsx)

    targets = []
    for r in rows:
        fid = str(r.get("fb_post_id") or "").strip()
        pid = str(r.get("post_id") or "").strip()
        if fid:  # bỏ qua bài chưa có fb_post_id
            targets.append((pid, fid))

    updated, failed = 0, 0
    session = requests.Session()
    for pid, fid in targets:
        try:
            eng = fetch_engagement(fid, token, session)
        except Exception as e:
            failed += 1
            print(f"  [skip] post_id={pid} fb_post_id={fid}: {_err(e)}")
            continue
        eng.update(fetch_insights_best_effort(fid, token, session))
        row = {
            "post_id": pid,
            "fb_post_id": fid,
            "fb_permalink": eng.get("fb_permalink", ""),
            "likes": eng.get("likes"),
            "comments": eng.get("comments"),
            "reactions": eng.get("reactions"),
            "shares": eng.get("shares"),
            "reach": eng.get("reach"),
            "impressions": eng.get("impressions"),
            "fetched_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        }
        # bỏ key None để không ghi đè bằng rỗng vô nghĩa (vẫn upsert đủ cột chính)
        row = {k: ("" if v is None else v) for k, v in row.items()}
        try:
            _upsert_engagement(campaign_xlsx, row)
        except Exception as e:
            failed += 1
            print(f"  [excel-fail] post_id={pid}: {_err(e)}")
            continue
        updated += 1
        print(f"  [ok] post_id={pid} likes={row['likes']} comments={row['comments']} "
              f"reactions={row['reactions']} shares={row['shares']} "
              f"reach={row['reach'] or '-'} impr={row['impressions'] or '-'}")

    print(f"Engagement: {updated} cập nhật / {failed} lỗi / "
          f"{len(targets)} bài có fb_post_id (tổng {len(rows)} Result rows).")
    return updated


def _err(e):
    if isinstance(e, requests.HTTPError) and e.response is not None:
        try:
            return str(e.response.json().get("error", {}).get("message", e))[:160]
        except Exception:
            return str(e)[:160]
    return f"{type(e).__name__}: {e}"[:160]


def main(argv=None):
    ap = argparse.ArgumentParser(description="Quét engagement FB -> sheet Engagement.")
    ap.add_argument("--campaign-xlsx", required=True)
    ap.add_argument("--fb-config", default=DEFAULT_FB_CONFIG)
    args = ap.parse_args(argv)

    if not os.path.isfile(args.campaign_xlsx):
        sys.exit(f"Không thấy campaign Excel: {args.campaign_xlsx}")

    run(args.campaign_xlsx, args.fb_config)
    print(f"OK {os.path.abspath(args.campaign_xlsx)}")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as e:  # in lỗi gọn + exit !=0
        print(f"ERROR {type(e).__name__}: {e}", file=sys.stderr)
        raise SystemExit(1)

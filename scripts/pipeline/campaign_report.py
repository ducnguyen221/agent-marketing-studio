# -*- coding: utf-8 -*-
r"""campaign_report.py — tổng hợp 1 campaign + APPEND báo cáo vào file .md.

  PY campaign_report.py --campaign-dir <CAMPAIGNS\NN_Ten> [--refresh-engagement] [--date YYYY-MM-DD]

- Tìm <NN_Ten>.xlsx + <NN_Ten>.md trong --campaign-dir (NN_Ten = tên folder).
- --refresh-engagement: gọi fb_engagement.py trước (subprocess) để cập nhật sheet
  Engagement (best-effort — lỗi không chặn báo cáo).
- Đọc Post + Result + Engagement (read-only, đọc header động để bền với việc
  tầng Excel đang được agent khác viết lại) -> in báo cáo từng bài + rollup.
- APPEND khối "## Báo cáo <ngày>" vào file .md. Ngày = --date nếu có, else now().

In dòng cuối: `OK <md path>`.
"""
from __future__ import annotations

import argparse
import datetime
import os
import subprocess
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
FB_ENGAGEMENT = os.path.join(_HERE, "fb_engagement.py")
DEFAULT_FB_CONFIG = os.environ.get("FB_CONFIG") or os.path.join(
    os.path.expanduser("~"), ".video", "tobi", "facebook_config.json")


# ---------------------------------------------------------------- đọc Excel (read-only, schema-tolerant)

def _read_sheet(xlsx_path, sheet):
    """list dict theo header thực tế. [] nếu thiếu file/sheet."""
    if not os.path.isfile(xlsx_path):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if sheet not in wb.sheetnames:
            return []
        rows = list(wb[sheet].iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec, empty = {}, True
        for i, col in enumerate(header):
            if not col:
                continue
            v = r[i] if i < len(r) else None
            if v not in (None, ""):
                empty = False
            rec[col] = v
        if not empty:
            out.append(rec)
    return out


def _by_post_id(rows):
    """index list dict theo post_id (str)."""
    idx = {}
    for r in rows:
        pid = str(r.get("post_id") or "").strip()
        if pid:
            idx[pid] = r
    return idx


def _num(v):
    """ép số an toàn cho rollup; non-numeric -> 0."""
    try:
        if v in (None, ""):
            return 0
        return int(float(v))
    except (TypeError, ValueError):
        return 0


# ---------------------------------------------------------------- build report

def build_report(campaign_dir, name, do_refresh, fb_config, report_date):
    xlsx = os.path.join(campaign_dir, f"{name}.xlsx")
    md = os.path.join(campaign_dir, f"{name}.md")
    notes = []

    if do_refresh:
        if os.path.isfile(xlsx):
            try:
                proc = subprocess.run(
                    [sys.executable, FB_ENGAGEMENT,
                     "--campaign-xlsx", xlsx, "--fb-config", fb_config],
                    capture_output=True, text=True, encoding="utf-8")
                if proc.returncode != 0:
                    tail = (proc.stderr or proc.stdout or "").strip().splitlines()
                    notes.append("refresh-engagement lỗi: "
                                 + (tail[-1] if tail else f"exit {proc.returncode}"))
                else:
                    notes.append("refresh-engagement OK")
            except Exception as e:
                notes.append(f"refresh-engagement exception: {type(e).__name__}: {e}")
        else:
            notes.append("refresh-engagement bỏ qua: chưa có file .xlsx")

    posts = _read_sheet(xlsx, "Post")
    results = _by_post_id(_read_sheet(xlsx, "Result"))
    engagement = _by_post_id(_read_sheet(xlsx, "Engagement"))

    lines = []
    lines.append(f"## Báo cáo {report_date}")
    lines.append("")
    for n in notes:
        lines.append(f"> _{n}_")
    if notes:
        lines.append("")

    total = len(posts)
    published = 0
    tot_likes = tot_comments = tot_shares = tot_reactions = 0

    if not posts:
        lines.append("_Chưa có bài nào trong sheet Post._")
    else:
        lines.append("| post_id | status | đã đăng | link | like | comment | share |")
        lines.append("|---|---|---|---|---|---|---|")
        for p in posts:
            pid = str(p.get("post_id") or "").strip()
            status = str(p.get("status") or "").strip() or "-"
            res = results.get(pid, {})
            eng = engagement.get(pid, {})
            fb_id = str(res.get("fb_post_id") or "").strip()
            pub_at = str(res.get("published_at") or "").strip()
            is_pub = bool(fb_id) or (str(res.get("status") or "").strip().lower() == "published")
            if is_pub:
                published += 1
            link = (str(eng.get("fb_permalink") or "").strip()
                    or str(res.get("fb_permalink") or "").strip()
                    or str(res.get("blog_url") or "").strip()
                    or str(res.get("youtube_url") or "").strip() or "-")
            likes = _num(eng.get("likes"))
            comments = _num(eng.get("comments"))
            shares = _num(eng.get("shares"))
            tot_likes += likes
            tot_comments += comments
            tot_shares += shares
            tot_reactions += _num(eng.get("reactions"))
            pub_mark = ("✓ " + pub_at) if is_pub and pub_at else ("✓" if is_pub else "—")
            link_cell = f"[link]({link})" if link != "-" else "—"
            lines.append(f"| {pid or '-'} | {status} | {pub_mark} | {link_cell} | "
                         f"{likes} | {comments} | {shares} |")

    lines.append("")
    lines.append("### Rollup")
    lines.append(f"- Tổng bài: **{total}** · đã đăng: **{published}**")
    lines.append(f"- Engagement: likes **{tot_likes}** · comments **{tot_comments}** "
                 f"· reactions **{tot_reactions}** · shares **{tot_shares}** "
                 f"· tổng tương tác **{tot_likes + tot_comments + tot_reactions + tot_shares}**")
    lines.append("")

    block = "\n".join(lines)

    # in báo cáo ra stdout
    print(block)

    # APPEND vào .md (tạo nếu chưa có)
    prefix = ""
    if os.path.isfile(md):
        with open(md, encoding="utf-8-sig") as f:
            existing = f.read()
        if existing and not existing.endswith("\n"):
            prefix = "\n"
        prefix += "\n"
    else:
        prefix = f"# {name}\n\n"
    with open(md, "a", encoding="utf-8") as f:
        f.write(prefix + block + "\n")

    return md


def main(argv=None):
    ap = argparse.ArgumentParser(description="Báo cáo campaign + append vào .md.")
    ap.add_argument("--campaign-dir", required=True)
    ap.add_argument("--refresh-engagement", action="store_true")
    ap.add_argument("--date", default="", help="YYYY-MM-DD (mặc định: hôm nay)")
    ap.add_argument("--fb-config", default=DEFAULT_FB_CONFIG)
    args = ap.parse_args(argv)

    campaign_dir = os.path.abspath(args.campaign_dir)
    if not os.path.isdir(campaign_dir):
        sys.exit(f"Không thấy campaign dir: {campaign_dir}")
    name = os.path.basename(campaign_dir.rstrip("\\/"))

    report_date = args.date.strip() or datetime.datetime.now().strftime("%Y-%m-%d")

    md = build_report(campaign_dir, name, args.refresh_engagement,
                      args.fb_config, report_date)
    print(f"OK {os.path.abspath(md)}")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    try:
        raise SystemExit(main())
    except SystemExit:
        raise
    except Exception as e:
        print(f"ERROR {type(e).__name__}: {e}", file=sys.stderr)
        raise SystemExit(1)

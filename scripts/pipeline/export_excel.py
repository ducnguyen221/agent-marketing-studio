#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""export_excel.py — xuất một chiến dịch ra .xlsx.

**Excel giờ là SẢN PHẨM XUẤT, không phải nguồn sự thật.** Nguồn duy nhất là `campaign.md`
(+ `research.md`/`publish.json` của từng bài). File này đi MỘT CHIỀU: Markdown → Excel.

Vì sao đổi: Excel là định dạng nhị phân — không diff được, không merge được, mở bằng Excel
là khoá file, và một agent ghi đè lúc người đang mở là mất trắng công cả hai bên. Markdown
diff được từng dòng, ai sửa gì thấy ngay trong git.

Vì sao vẫn xuất: người ta cần lọc, xoay, gửi cho người không dùng git, và làm báo cáo.

Giữ NGUYÊN bộ cột của `CAMPAIGN_TEMPLATE.xlsx` cũ (3 sheet Campaign · Content · Post) để
biểu mẫu và thói quen cũ không phải học lại. Trường nào Markdown chưa có thì để TRỐNG —
ô rỗng nói "chưa biết", đừng bịa số vào cho đầy bảng.

CLI:
  python export_excel.py --campaign <đường/dẫn>            # → <thư mục>/<id>.xlsx
  python export_excel.py --station ~/.marketing            # xuất mọi chiến dịch
  python export_excel.py --campaign <đd> --out D:/bao-cao.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "lib"))
import md_io  # noqa: E402
import studio_paths as SP  # noqa: E402

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:      # noqa: BLE001
    openpyxl = None

# Bộ cột GIỮ NGUYÊN từ CAMPAIGN_TEMPLATE.xlsx v3 — đổi thứ tự là biểu mẫu cũ hỏng.
COT_CONTENT = ["content_id", "content_name", "content_pillar", "funnel_stage",
               "content_angle", "priority", "content_goal", "audience_profile",
               "core_brief", "key_sources", "target_keyword", "creative_direction",
               "constraints", "content_relationship", "audio", "video", "short",
               "status", "approved_date", "schedule_date", "published_date",
               "folder_path", "notes"]
COT_POST = ["post_id", "content_id", "channel", "post_format", "post_role", "post_content",
            "quality_check", "agent_status", "review_status", "review_feedback",
            "post_status", "publish_plan", "publish_status", "publish_link",
            "target_view", "target_interaction", "updated_at", "asset_ref",
            "actual_view", "actual_interaction", "actual_reaction", "actual_comment",
            "actual_share", "actual_click", "actual_reach", "metric_updated_at"]

# Cột bảng Content trong campaign.md → cột sheet Content. Cột nào không có ở đây là
# trường Excel-only, để trống.
ANH_XA = {"content_id": "content_id", "content_name": "content_name",
          "pillar": "content_pillar", "funnel": "funnel_stage", "angle": "content_angle",
          "priority": "priority", "status": "status", "g1": "approved_date",
          "schedule": "schedule_date", "published": "published_date",
          "folder": "folder_path"}

XANH = PatternFill("solid", fgColor="1F3864") if openpyxl else None


def _dong_content(cam_dir: Path, dong_md: list) -> list[dict]:
    """Một dòng Content = dòng trong campaign.md, làm giàu bằng research.md của bài."""
    ra = []
    for d in dong_md:
        o = {k: "" for k in COT_CONTENT}
        for md_k, xl_k in ANH_XA.items():
            if d.get(md_k):
                o[xl_k] = d[md_k]
        thu_muc = (d.get("folder") or "").strip("./")
        r = cam_dir / thu_muc / "research.md"
        if thu_muc and r.is_file():
            fm, _ = md_io.read_fm(r)
            for xl_k, fm_k in (("content_goal", "content_goal"),
                               ("audience_profile", "audience_profile"),
                               ("core_brief", "core_brief"),
                               ("target_keyword", "target_keyword"),
                               ("creative_direction", "creative_direction"),
                               ("constraints", "constraints"),
                               ("audio", "audio"), ("video", "video"), ("short", "short")):
                v = fm.get(fm_k)
                if v not in (None, "", []):
                    o[xl_k] = ", ".join(v) if isinstance(v, list) else str(v)
        ra.append(o)
    return ra


def _dong_post(cam_dir: Path, dong_md: list) -> list[dict]:
    """Một dòng Post = một mục trong `publish.json` của bài. Chưa đăng thì chưa có dòng."""
    ra = []
    for d in dong_md:
        thu_muc = (d.get("folder") or "").strip("./")
        pj_p = cam_dir / thu_muc / "publish.json"
        if not (thu_muc and pj_p.is_file()):
            continue
        pj = json.loads(pj_p.read_text(encoding="utf-8"))
        for p in pj.get("posts", []):
            o = {k: "" for k in COT_POST}
            pub, rev = p.get("publish", {}), p.get("review", {})
            # Ba lỗi đọc-nhầm-khoá đã sửa 05/09, đều làm ô im lặng rỗng hoặc sai nội dung:
            #   · `publish_plan` nằm ở `publish.plan`, không phải ở gốc post;
            #   · `updated_at` của review là `approved_at`, không phải `at`;
            #   · `review_feedback` là `review.feedback` (góp ý), còn `review.note` là CÂU
            #     DUYỆT nguyên văn — lấy nhầm thì cột góp ý in ra lời chấp thuận.
            # Và năm cột trạng thái trước đây không được ánh xạ gì cả, nên mọi file .xlsx
            # xuất ra đều trống ở đó dù publish.json ghi rõ passed/completed/published —
            # người nhận Excel tưởng bài chưa qua cổng kỹ thuật.
            o.update({
                "post_id": p.get("post_id", ""), "content_id": pj.get("post_id", ""),
                "channel": p.get("channel", ""), "post_format": p.get("post_format", ""),
                "post_role": p.get("post_role", ""), "post_content": p.get("post_content", ""),
                "asset_ref": p.get("asset_ref", ""),
                "quality_check": p.get("quality_check", ""),
                "agent_status": p.get("agent_status", ""),
                "post_status": p.get("post_status", ""),
                "review_status": rev.get("status", ""),
                "review_feedback": rev.get("feedback", ""),
                "publish_status": pub.get("status", ""), "publish_link": pub.get("link", ""),
                "publish_plan": pub.get("plan", ""),
                "updated_at": p.get("updated_at", "") or pub.get("at", "")
                              or rev.get("approved_at", ""),
            })
            # Khoá THẬT trong publish.json là `actual` và `target` — `register_publish
            # metrics` ghi vào đó. Đọc nhầm tên khoá thì mọi ô số liệu im lặng rỗng.
            for ten in ("view", "interaction", "reaction", "comment", "share", "click", "reach"):
                v = (p.get("actual") or {}).get(ten)
                if v is not None:
                    o["actual_" + ten] = v
            for ten in ("view", "interaction"):
                v = (p.get("target") or {}).get(ten)
                if v is not None:
                    o["target_" + ten] = v
            if (p.get("actual") or {}).get("updated_at"):
                o["metric_updated_at"] = p["actual"]["updated_at"]
            ra.append(o)
    return ra


def _sheet(wb, ten: str, cot: list[str], dong: list[dict]):
    ws = wb.create_sheet(ten)
    ws.append(cot)
    for c in range(1, len(cot) + 1):
        o = ws.cell(1, c)
        o.font = Font(bold=True, color="FFFFFF")
        o.fill = XANH
        o.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = max(12, min(38, len(cot[c - 1]) + 8))
    for d in dong:
        ws.append([d.get(k, "") for k in cot])
    ws.freeze_panes = "A2"
    if dong:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cot))}{len(dong) + 1}"
    return ws


def xuat(cam_dir: Path, out: Path | None = None) -> Path:
    fm, body = md_io.read_fm(cam_dir / "campaign.md")
    _, dong_md = md_io.read_table(body, "CONTENT")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Campaign")
    ws.append(["field", "value", "agent_instruction"])
    for c in (1, 2, 3):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = XANH
    # Dict lồng (kpi_target…) trải thành từng dòng `kpi_target.blog`. Nhét cả object vào
    # một ô thì openpyxl từ chối, mà có nhét được cũng không ai lọc được.
    for k, v in fm.items():
        if isinstance(v, dict):
            for k2, v2 in v.items():
                ws.append([f"{k}.{k2}", v2 if v2 is not None else "", ""])
        else:
            ws.append([k, ", ".join(map(str, v)) if isinstance(v, list) else v, ""])
    ws.append(["", "", ""])
    # Đường dẫn TƯƠNG ĐỐI (kênh/chiến_dịch/campaign.md). Đường tuyệt đối mang theo tên
    # người dùng và tên máy — file .xlsx này được gửi đi và có bản nằm trong repo công khai.
    neo = f"{cam_dir.parent.name}/{cam_dir.name}/campaign.md"
    ws.append(["⚠️ NGUỒN SỰ THẬT", neo,
               "File Excel này là BẢN XUẤT. Sửa ở đây KHÔNG quay ngược về Markdown — "
               "sửa campaign.md rồi xuất lại."])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="C00000")
    for c, w in ((1, 26), (2, 62), (3, 70)):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"

    _sheet(wb, "Content", COT_CONTENT, _dong_content(cam_dir, dong_md))
    _sheet(wb, "Post", COT_POST, _dong_post(cam_dir, dong_md))

    dich = Path(out) if out else cam_dir / f"{fm.get('id', cam_dir.name)}.xlsx"
    dich.parent.mkdir(parents=True, exist_ok=True)
    # openpyxl không ghi nguyên tử được; ghi tạm rồi đổi tên, để người đang mở file cũ
    # không gặp một file .xlsx cụt.
    tmp = dich.with_suffix(".xlsx.tmp")
    wb.save(tmp)
    tmp.replace(dich)
    return dich


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Xuất chiến dịch ra .xlsx (một chiều từ Markdown).")
    ap.add_argument("--campaign", default=None)
    ap.add_argument("--station", default=None)
    ap.add_argument("--out", default=None, help="chỉ dùng với --campaign")
    a = ap.parse_args(argv)

    if openpyxl is None:
        sys.stderr.write("cần openpyxl: pip install openpyxl\n")
        return 3
    if a.out and not a.campaign:
        sys.stderr.write("--out chỉ đi cùng --campaign\n")
        return 2

    if a.campaign:
        cam = Path(a.campaign).resolve()
        if not (cam / "campaign.md").is_file():
            sys.stderr.write(f"không thấy campaign.md trong {cam}\n")
            return 2
        print(f"  {xuat(cam, Path(a.out) if a.out else None)}")
        return 0

    station = SP.root(a.station).resolve()
    n = 0
    for k in SP.channels(station):
        for c in sorted(Path(k["dir"]).iterdir()) if Path(k["dir"]).is_dir() else []:
            if (c / "campaign.md").is_file():
                print(f"  {xuat(c)}")
                n += 1
    if not n:
        sys.stderr.write("không có chiến dịch nào để xuất\n")
        return 1
    print(f"  {n} file .xlsx")
    return 0


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass
    raise SystemExit(main())
